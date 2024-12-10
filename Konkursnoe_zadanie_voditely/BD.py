import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect("drivers.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS drivers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    guid TEXT NOT NULL,
    last_name TEXT NOT NULL,
    first_name TEXT NOT NULL,
    middle_name TEXT,
    passport TEXT NOT NULL,
    registration_city TEXT,
    registration_address TEXT,
    living_city TEXT,
    living_address TEXT,
    workplace TEXT,
    position TEXT,
    phone TEXT,
    email TEXT,
    notes TEXT
);
""")
conn.commit()

try:
    cursor.execute("ALTER TABLE drivers ADD COLUMN photo_path TEXT;")
except sqlite3.OperationalError:
    pass

wb = load_workbook("drivers.xlsx")
sheet = wb.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    cursor.execute("""
    INSERT INTO drivers (guid, last_name, first_name, middle_name, passport, registration_city, 
                         registration_address, living_city, living_address, workplace, position, phone, email, notes)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, row)

conn.commit()
conn.close()
